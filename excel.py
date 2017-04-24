import openpyxl
import os
import tablib
import xlrd


def generate_workbook(workbook_file_path):

    wb = openpyxl.Workbook()

    ws = wb.create_sheet("Test Cases")
    fill = openpyxl.styles.PatternFill("solid", fgColor="DDDDDD")
    ws.cell(row=1, column=1, value="Test Case ID").fill = fill
    ws.cell(row=1, column=2, value="Active").fill = fill
    ws.cell(row=1, column=2, value="Test Case Name").fill = fill
    ws.cell(row=1, column=2, value="Test Case Description").fill = fill

    # remove default sheet
    del wb["Sheet"]

    wb.save(workbook_file_path)

def get_workbook_metadata(workbook_file_path):
    pass
    

def load_workbook(workbook_file_path):
    #dataset = tablib.Dataset().xlsx = open(workbook_file_path,"rb").read()
    #databook = tablib.Databook().load("xlsx", open(workbook_file_path).read())
    #print(dataset[0])
    
    #wb = openpyxl.load_workbook(workbook_file_path)
    #print(wb.sheetnames)
    
    wb = xlrd.open_workbook(workbook_file_path)
    return wb

    
