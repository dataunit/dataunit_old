import unittest
import openpyxl
import os
import excel

class TestExcel(unittest.TestCase):

    def test_generate_workbook(self):

        wb_path = '/Users/dmoore/Documents/projects/dataunit/wb.xlsx'

        # delete file if it exists
        if os.path.isfile(wb_path):
            os.remove(wb_path)

        # call function to create the workbook
        excel.generate_workbook(wb_path)
        
        # verify file exists
        self.assertTrue(os.path.isfile(wb_path))
        
        # attempt to load workbook
        wb = openpyxl.load_workbook(wb_path)

        # assert has a sheet named Test Cases
        self.assertIn("Test Cases", wb.sheetnames, "Sheet 'Test Cases' not found in Workbook")
        
        # assert that the sheet has column headers
        ws = wb["Test Cases"]
        self.assertEqual(ws.cell(row=1,column=1).value, "Test Case ID", "Test Case ID column not found")

        # delete file
        os.remove(wb_path)


    def test_load_workbook(self):
        wb_path = '/Users/dmoore/Documents/projects/dataunit/test_case_workbook.v2.xlsx'
        wb = excel.load_workbook(wb_path)
        self.assertEqual(wb.sheet_by_index(0).name,"Test Cases")

    
