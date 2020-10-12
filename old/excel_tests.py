import unittest
import openpyxl
import os
import excel
import sqlite3
import pandas as pd

class TestExcel(unittest.TestCase):

    @unittest.skip("not using this")
    def test_generate_workbook(self):

        wb_path = '/Users/dmoore/Documents/projects/dataunit/wb.xlsx'

        # delete file if it exists
        if os.path.isfile(wb_path):
            os.remove(wb_path)

        # call function to create the workbook
        excel.generate_workbook(wb_path)
        
        # verify file exists
        self.assertTrue(os.path.isfile(wb_path))
        
        # attempt to load workbook
        wb = openpyxl.load_workbook(wb_path)

        # assert has a sheet named Test Cases
        self.assertIn("Test Cases", wb.sheetnames, "Sheet 'Test Cases' not found in Workbook")
        
        # assert that the sheet has column headers
        ws = wb["Test Cases"]
        self.assertEqual(ws.cell(row=1,column=1).value, "Test Case ID", "Test Case ID column not found")

        # delete file
        os.remove(wb_path)


    def test_load_workbook(self):
        wb_path = '/Users/dmoore/Documents/projects/dataunit/test_case_workbook.v3.xlsx'
        wb = excel.load_workbook(wb_path)
        self.assertEqual(wb.sheet_by_index(0).name,"Tests")

    #@unittest.skip("not using this")
    def test_load_workbook_to_db(self):
        wb_path = '/Users/dmoore/Documents/projects/dataunit/test_case_workbook.v3.xlsx'
        db_path = '/Users/dmoore/Documents/projects/dataunit/test_case_workbook.v3.db'
        conn = excel.load_workbook_to_db(wb_path,db_path)
        cursor = conn.cursor()
        #print(cursor.execute("select name from sqlite_master where type = 'table'").fetchall())

    def test_get_column_code_mapper(self):
        wb_path = '/Users/dmoore/Documents/projects/dataunit/test_case_workbook.v3.xlsx'
        wb = pd.ExcelFile(wb_path)
        df = wb.parse(excel.SHEETNAME_WORKBOOK_COLUMN_METADATA)
        get_column_code = excel.get_column_code_mapper(df)

        column_code = get_column_code("Tests","Test ID")
        self.assertEqual(column_code,"test_id")

        column_code = get_column_code("Test Commands","Command Sequence Number")
        self.assertEqual(column_code,"command_sequence_number")

    def test_get_sheet_code_mapper(self):
        wb_path = '/Users/dmoore/Documents/projects/dataunit/test_case_workbook.v3.xlsx'
        wb = pd.ExcelFile(wb_path)
        df = wb.parse(excel.SHEETNAME_WORKBOOK_SHEET_METADATA)
        get_sheet_code = excel.get_sheet_code_mapper(df)

        sheet_code = get_sheet_code("Tests")
        self.assertEqual(sheet_code,"tests")

        sheet_code = get_sheet_code("Test Commands")
        self.assertEqual(sheet_code,"test_commands")
        
   
    #TODO: write test for sheet missing exception