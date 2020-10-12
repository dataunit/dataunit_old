import openpyxl
import os
import tablib
import xlrd
import pandas as pd
import sqlite3
from exception import SheetMissingError

SHEETNAME_WORKBOOK_SHEET_METADATA = "_workbook_sheet_metadata"
SHEETNAME_WORKBOOK_COLUMN_METADATA = "_workbook_column_metadata"


def generate_workbook(workbook_file_path):

    wb = openpyxl.Workbook()

    ws = wb.create_sheet("Test Cases")
    fill = openpyxl.styles.PatternFill("solid", fgColor="DDDDDD")
    ws.cell(row=1, column=1, value="Test Case ID").fill = fill
    ws.cell(row=1, column=2, value="Active").fill = fill
    ws.cell(row=1, column=2, value="Test Case Name").fill = fill
    ws.cell(row=1, column=2, value="Test Case Description").fill = fill

    # remove default sheet
    del wb["Sheet"]

    wb.save(workbook_file_path)

def get_workbook_metadata(workbook_file_path):
    pass


def load_workbook(workbook_file_path):
    #dataset = tablib.Dataset().xlsx = open(workbook_file_path,"rb").read()
    #databook = tablib.Databook().load("xlsx", open(workbook_file_path).read())
    #print(dataset[0])
    
    #wb = openpyxl.load_workbook(workbook_file_path)
    #print(wb.sheetnames)
    
    wb = xlrd.open_workbook(workbook_file_path)
    return wb


def load_workbook_to_db(workbook_file_path,db_file_path):
    wb = pd.ExcelFile(workbook_file_path)
    conn = sqlite3.connect(db_file_path)

    #build mapping functions from metadata sheets
    sheet_name = SHEETNAME_WORKBOOK_SHEET_METADATA
    if sheet_name not in wb.sheet_names:
        raise SheetMissingError("Sheet %s not found in workbook %s" % (sheet_name, os.path.basename(workbook_file_path)))
    df = wb.parse(sheet_name)
    get_sheet_code = get_sheet_code_mapper(df)
    print(df.to_string())

    sheet_name = SHEETNAME_WORKBOOK_COLUMN_METADATA
    if sheet_name not in wb.sheet_names:
        raise SheetMissingError("Sheet %s not found in workbook %s" % (sheet_name, os.path.basename(workbook_file_path)))
    df = wb.parse(sheet_name)
    get_column_code = get_column_code_mapper(df)    


    for sheet in wb.sheet_names:
        df = wb.parse(sheet)
        sheet_code = get_sheet_code(sheet)

        get_column_sheet_code = lambda column_name: get_column_code(sheet,column_name)

        df.rename(columns=get_column_sheet_code, inplace=True)

        df.to_sql(sheet_code,conn,if_exists="replace")

    return conn

def get_column_code_mapper(df):
    "builds a function that returns a sheet/column mapping function from a dataframe"

    # sheet_name = excel.SHEETNAME_WORKBOOK_COLUMN_METADATA
    # if sheet_name not in wb.sheet_names:  
    #     raise Exception("Sheet %s not found in workbook %s" % sheet_name, wb.Name)
    my_dict = {}

    for index, row in df.iterrows():
        my_dict[row.sheet_name + "." + row.column_name.strip()] = row.column_code_name

    return lambda sheet_name, column_name: my_dict.get(sheet_name + "." + column_name.strip(),column_name)

def get_sheet_code_mapper(df):
    "builds a function that returns a sheet mapping function from a dataframe"
    my_dict = {}

    for index, row in df.iterrows():
        my_dict[row.sheet_name] = row.sheet_code_name

    return lambda sheet_name: my_dict.get(sheet_name,sheet_name)